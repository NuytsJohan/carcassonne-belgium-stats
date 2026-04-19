"""Export Belgian national friendly stats into the BGA Excel template format.

Produces a 4-sheet workbook mirroring `data/raw/Belgium Friendlies.xlsx`:
  - matches_csv: one row per country-vs-country friendly match
  - duels_csv:   one row per individual player-vs-player Bo3 duel
  - profiles_csv: one row per distinct BGA player involved
  - associations: country name -> ISO3 (copied from the template)
"""

from __future__ import annotations

from pathlib import Path
from datetime import datetime, time

import duckdb
import openpyxl


ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "carcassonne.duckdb"
TEMPLATE_PATH = ROOT / "data" / "raw" / "Belgium Friendlies.xlsx"
OUTPUT_PATH = ROOT / "data" / "raw" / "Belgium Friendlies - Filled.xlsx"

TOURNAMENT_ID_IN_DB = 1          # 'Friendlies' row in tournaments table
BGA_TOURNAMENT_ID = "Belgium-Friendlies"

# Player.country (ISO2) -> ISO3 used by the BGA template.
ISO2_TO_ISO3 = {
    "AR": "ARG", "AT": "AUT", "AU": "AUS", "BE": "BEL", "BG": "BGR",
    "BR": "BRA", "CA": "CAN", "CH": "CHE", "CL": "CHL", "CN": "CHN",
    "CO": "COL", "CR": "CRI", "CU": "CUB", "CZ": "CZE", "DE": "DEU",
    "DK": "DNK", "EC": "ECU", "EE": "EST", "EG": "EGY", "ES": "ESP",
    "FI": "FIN", "FR": "FRA", "GB": "GBR", "GR": "GRC", "GT": "GTM",
    "HK": "HKG", "HR": "HRV", "HU": "HUN", "ID": "IDN", "IL": "ISR",
    "IN": "IND", "IS": "ISL", "IT": "ITA", "JP": "JPN", "KR": "KOR",
    "KZ": "KAZ", "LT": "LTU", "LU": "LUX", "LV": "LVA", "MD": "MDA",
    "MX": "MEX", "MY": "MYS", "NL": "NLD", "NO": "NOR", "PE": "PER",
    "PL": "POL", "PT": "PRT", "RO": "ROU", "RS": "SRB", "SE": "SWE",
    "SG": "SGP", "SK": "SVK", "TH": "THA", "TR": "TUR", "TW": "TWN",
    "UA": "UKR", "US": "USA", "UY": "URY", "VN": "VNM",
}

# Opponent country (as stored in nations_competition_duels.opponent_country) -> ISO3.
OPPONENT_TO_ISO3 = {
    "Argentina": "ARG", "Australia": "AUS", "Brazil": "BRA",
    "Catalonia": "CAT", "China": "CHN", "Colombia": "COL",
    "Croatia": "HRV", "Finland": "FIN", "France": "FRA",
    "Germany": "DEU", "Greece": "GRC", "Hong Kong": "HKG",
    "Hungary": "HUN", "Italy": "ITA", "Japan": "JPN",
    "Latvia": "LVA", "Poland": "POL", "Romania": "ROU",
    "Spain": "ESP", "Sweden": "SWE", "The Netherlands": "NLD",
    "UK": "GBR", "USA": "USA", "Ukraine": "UKR",
}


def country_to_iso3(iso2: str | None) -> str:
    if not iso2:
        return "UNK"
    return ISO2_TO_ISO3.get(iso2.upper(), iso2.upper())


def fetch_matches(con: duckdb.DuckDBPyConnection) -> list[dict]:
    rows = con.execute(
        """
        SELECT d.id                                               AS duel_id,
               d.stage                                            AS stage,
               d.date_played                                      AS date_played,
               d.opponent_country                                 AS opponent,
               COUNT(nm.id)                                       AS num_duels,
               SUM(CASE WHEN nm.result = 'W' THEN 1 ELSE 0 END)   AS dw_bel,
               SUM(CASE WHEN nm.result = 'L' THEN 1 ELSE 0 END)   AS dw_opp
        FROM nations_competition_duels d
        JOIN nations_matches nm ON nm.duel_id = d.id
        WHERE d.tournament_id = ?
        GROUP BY d.id, d.stage, d.date_played, d.opponent_country
        ORDER BY d.date_played, d.id
        """,
        [TOURNAMENT_ID_IN_DB],
    ).fetchall()
    return [
        {
            "duel_id": r[0],
            "stage": r[1],
            "date_played": r[2],
            "opponent": r[3],
            "num_duels": int(r[4] or 0),
            "dw_bel": int(r[5] or 0),
            "dw_opp": int(r[6] or 0),
        }
        for r in rows
    ]


def fetch_duels(con: duckdb.DuckDBPyConnection) -> list[dict]:
    """Fetch one row per nations_matches with games-won counts from game data."""
    rows = con.execute(
        """
        WITH fm AS (
            SELECT nm.id                AS nm_id,
                   nm.duel_id           AS duel_id,
                   nm.player_id         AS bel_player_id,
                   nm.opponent_player_id AS opp_player_id,
                   nm.result            AS result,
                   nm.notes             AS notes,
                   nm.score_belgium     AS score_bel,
                   nm.score_opponent    AS score_opp,
                   unnest([nm.game_id_1, nm.game_id_2, nm.game_id_3,
                           nm.game_id_4, nm.game_id_5]) AS game_id
            FROM nations_matches nm
            JOIN nations_competition_duels d ON d.id = nm.duel_id
            WHERE d.tournament_id = ?
        ),
        winners AS (
            SELECT fm.nm_id, fm.bel_player_id, fm.opp_player_id,
                   g.played_at, gp.player_id AS winner_id
            FROM fm
            JOIN games g ON g.id = fm.game_id
            JOIN game_players gp ON gp.game_id = g.id AND gp.rank = 1
            WHERE fm.game_id IS NOT NULL
        ),
        agg AS (
            SELECT nm_id,
                   MIN(played_at) AS first_played_at,
                   SUM(CASE WHEN winner_id = bel_player_id THEN 1 ELSE 0 END) AS gw_bel,
                   SUM(CASE WHEN winner_id = opp_player_id THEN 1 ELSE 0 END) AS gw_opp
            FROM winners
            GROUP BY nm_id
        )
        SELECT nm.id,
               nm.duel_id,
               nm.player_id,
               nm.opponent_player_id,
               nm.result,
               nm.notes,
               nm.score_belgium,
               nm.score_opponent,
               d.date_played,
               agg.first_played_at,
               agg.gw_bel,
               agg.gw_opp
        FROM nations_matches nm
        JOIN nations_competition_duels d ON d.id = nm.duel_id
        LEFT JOIN agg ON agg.nm_id = nm.id
        WHERE d.tournament_id = ?
        ORDER BY d.date_played, d.id, nm.id
        """,
        [TOURNAMENT_ID_IN_DB, TOURNAMENT_ID_IN_DB],
    ).fetchall()

    duels: list[dict] = []
    for r in rows:
        (nm_id, duel_id, bel_pid, opp_pid, result, notes,
         score_bel, score_opp, date_played, first_played_at, gw_bel, gw_opp) = r

        if gw_bel is None and gw_opp is None:
            # No game records stored: fall back from result/score.
            if (score_bel or 0) == 0 and (score_opp or 0) == 0:
                gw_bel, gw_opp = 0, 0
            elif result == "W":
                gw_bel, gw_opp = 2, 0
            elif result == "L":
                gw_bel, gw_opp = 0, 2
            else:
                gw_bel, gw_opp = 0, 0

        played_at = first_played_at or (
            datetime.combine(date_played, time(18, 0)) if date_played else None
        )
        duels.append({
            "nm_id": nm_id,
            "duel_id": duel_id,
            "bel_player_id": bel_pid,
            "opp_player_id": opp_pid,
            "played_at": played_at,
            "date_played": date_played,
            "gw_bel": int(gw_bel or 0),
            "gw_opp": int(gw_opp or 0),
            "notes": notes,
        })
    return duels


def fetch_profiles(con: duckdb.DuckDBPyConnection) -> list[dict]:
    rows = con.execute(
        """
        WITH friendly_players AS (
            SELECT DISTINCT nm.player_id AS pid
            FROM nations_matches nm
            JOIN nations_competition_duels d ON d.id = nm.duel_id
            WHERE d.tournament_id = ?
            UNION
            SELECT DISTINCT nm.opponent_player_id
            FROM nations_matches nm
            JOIN nations_competition_duels d ON d.id = nm.duel_id
            WHERE d.tournament_id = ?
        )
        SELECT p.id, p.name, p.bga_player_id, p.country
        FROM friendly_players f
        JOIN players p ON p.id = f.pid
        ORDER BY p.country, p.name
        """,
        [TOURNAMENT_ID_IN_DB, TOURNAMENT_ID_IN_DB],
    ).fetchall()
    return [
        {
            "player_id": r[0],
            "name": r[1],
            "bga_id": r[2],
            "country": r[3],
        }
        for r in rows
    ]


def load_associations(template_path: Path) -> list[tuple[str, str]]:
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb["associations"]
    return [(row[0], row[1]) for row in ws.iter_rows(values_only=True) if row and row[0]]


def build_workbook(
    matches: list[dict],
    duels: list[dict],
    profiles: list[dict],
    associations: list[tuple[str, str]],
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    # matches_csv
    ws = wb.active
    ws.title = "matches_csv"
    ws.append([
        "id", "tournament_id", "time_utc", "team_1", "team_2",
        "number_of_duels", "dw1", "dw2", "gw1", "gw2",
    ])
    # Per-match gw totals come from the duels; index by duel_id.
    gw_by_duel: dict[int, tuple[int, int]] = {}
    for d in duels:
        cur = gw_by_duel.get(d["duel_id"], (0, 0))
        gw_by_duel[d["duel_id"]] = (cur[0] + d["gw_bel"], cur[1] + d["gw_opp"])

    match_ids: dict[int, str] = {}
    for i, m in enumerate(matches, start=1):
        match_id = f"M{i}"
        match_ids[m["duel_id"]] = match_id
        gw_bel, gw_opp = gw_by_duel.get(m["duel_id"], (0, 0))
        time_utc = (
            datetime.combine(m["date_played"], time(18, 0))
            if m["date_played"] else None
        )
        ws.append([
            match_id,
            BGA_TOURNAMENT_ID,
            time_utc,
            "BEL",
            OPPONENT_TO_ISO3.get(m["opponent"], "UNK"),
            m["num_duels"],
            m["dw_bel"],
            m["dw_opp"],
            gw_bel,
            gw_opp,
        ])

    # duels_csv
    ws = wb.create_sheet("duels_csv")
    ws.append([
        "tournament_id", "match_id", "duel_number", "duel_format",
        "time_utc", "custom_time", "player_1_id", "player_2_id",
        "dw1", "dw2",
    ])
    # Group duels by their parent match to number them sequentially.
    duel_counter: dict[int, int] = {}
    pid_to_bga: dict[int, str | None] = {p["player_id"]: p["bga_id"] for p in profiles}
    for d in duels:
        duel_counter[d["duel_id"]] = duel_counter.get(d["duel_id"], 0) + 1
        duel_number = duel_counter[d["duel_id"]]
        match_id = match_ids.get(d["duel_id"])
        ws.append([
            BGA_TOURNAMENT_ID,
            match_id,
            duel_number,
            "Bo3",
            d["played_at"],
            0,
            pid_to_bga.get(d["bel_player_id"]),
            pid_to_bga.get(d["opp_player_id"]),
            d["gw_bel"],
            d["gw_opp"],
        ])

    # profiles_csv
    ws = wb.create_sheet("profiles_csv")
    ws.append(["id", "bga_nickname", "association"])
    for p in profiles:
        if not p["bga_id"]:
            continue
        ws.append([p["bga_id"], p["name"], country_to_iso3(p["country"])])

    # associations
    ws = wb.create_sheet("associations")
    for name, code in associations:
        ws.append([name, code])

    return wb


def main() -> None:
    con = duckdb.connect(str(DB_PATH), read_only=True)
    try:
        matches = fetch_matches(con)
        duels = fetch_duels(con)
        profiles = fetch_profiles(con)
    finally:
        con.close()

    associations = load_associations(TEMPLATE_PATH)
    wb = build_workbook(matches, duels, profiles, associations)
    wb.save(OUTPUT_PATH)

    print(f"Wrote {OUTPUT_PATH}")
    print(f"  matches:  {len(matches)}")
    print(f"  duels:    {len(duels)}")
    print(f"  profiles: {len([p for p in profiles if p['bga_id']])}")


if __name__ == "__main__":
    main()
