"""Import World Carcassonne Championship (WCC) results.

Source: data/raw/carcassonne_wc_results.xlsx
- Sheet 'WC Results': top 4 per year + metadata.
- Sheet 'Deelnemersgroei': per-year opmerkingen -> tournaments.notes.
- Sheet 'Belgische WK deelnames': Belgian participants (rank + note + BGA id).

Creates one tournament (type=WCC, name='WK {year}') per edition held.
The 2020 edition (COVID cancellation) is skipped.
"""
from pathlib import Path

import duckdb
import openpyxl

ROOT = Path(__file__).parents[1]
DB_PATH = ROOT / "data" / "carcassonne.duckdb"
XLSX = ROOT / "data" / "raw" / "carcassonne_wc_results.xlsx"

TOURNAMENT_ID_START = 100
YEARS_SKIP = {2020}

COUNTRY_CODES = {
    "Duitsland": "DE", "Nederland": "NL", "Tsjechië": "CZ", "Oostenrijk": "AT",
    "Taiwan": "TW", "Finland": "FI", "Noorwegen": "NO", "Australië": "AU",
    "Slowakije": "SK", "Japan": "JP", "Griekenland": "GR", "Letland": "LV",
    "Polen": "PL", "Portugal": "PT", "Brazilië": "BR", "Rusland": "RU",
    "Italië": "IT", "Roemenië": "RO", "Estland": "EE", "België": "BE",
    "VK": "GB", "Catalonië": "ES", "Hongarije": "HU", "China": "CN",
    "Uruguay": "UY",
}


def parse_participants(v):
    if v is None:
        return None
    if isinstance(v, int):
        return v if v > 0 else None
    s = str(v).strip().lstrip("~")
    try:
        n = int(s)
        return n if n > 0 else None
    except ValueError:
        return None


def country_code(dutch_name):
    if dutch_name is None:
        return None
    code = COUNTRY_CODES.get(dutch_name.strip())
    if code is None:
        raise KeyError(f"Unknown country: {dutch_name!r}")
    return code


def find_or_create_player(conn, name, country, bga_id=None):
    """Look up by name_nl / name (case-insensitive). Creates if missing.
    For every country: match either `name` or `name_nl` — foreign players
    imported via BGA/BCOC typically store the BGA handle in `name` and the
    real-world name in `name_nl`, so a lookup by the xlsx-supplied real name
    must check `name_nl` too.
    """
    row = conn.execute(
        "SELECT id FROM players "
        "WHERE LOWER(name) = LOWER(?) OR LOWER(name_nl) = LOWER(?)",
        [name, name],
    ).fetchone()

    if row:
        pid = row[0]
        if bga_id:
            existing = conn.execute(
                "SELECT bga_player_id FROM players WHERE id = ?", [pid]
            ).fetchone()[0]
            if not existing:
                conn.execute(
                    "UPDATE players SET bga_player_id = ? WHERE id = ?",
                    [bga_id, pid],
                )
        return pid, False

    if country == "BE":
        pid = conn.execute(
            "INSERT INTO players (name, name_nl, country, bga_player_id) "
            "VALUES (?, ?, 'BE', ?) RETURNING id",
            [name, name, bga_id],
        ).fetchone()[0]
    else:
        pid = conn.execute(
            "INSERT INTO players (name, country, bga_player_id) "
            "VALUES (?, ?, ?) RETURNING id",
            [name, country, bga_id],
        ).fetchone()[0]
    return pid, True


def load_workbook_data():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # Deelnemersgroei -> year -> opmerkingen
    notes_by_year = {}
    ws = wb["Deelnemersgroei"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        year = row[0]
        opm = row[3]
        if isinstance(year, int) and opm:
            notes_by_year[year] = opm

    # WC Results -> editions
    editions = []
    ws = wb["WC Results"]
    for row in list(ws.iter_rows(values_only=True))[1:]:
        year = row[0]
        if not isinstance(year, int) or year in YEARS_SKIP:
            continue
        champion = row[2]
        if not champion or str(champion).upper().startswith("AFGELAST"):
            continue
        editions.append({
            "year": year,
            "edition_no": row[1],
            "top4": [
                (1, row[2], row[3]),
                (2, row[4], row[5]),
                (3, row[6], row[7]),
                (4, row[8], row[9]),
            ],
            "participants_count": parse_participants(row[10]),
            "location": row[11],
        })

    # Belgische deelnames -> year -> list of (rank, name, bga_id, note)
    be_by_year = {}
    ws = wb["Belgische WK deelnames"]
    for row in list(ws.iter_rows(values_only=True))[1:]:
        year, name, bga_id, pos, _total, opm = row
        if not isinstance(year, int):
            continue
        be_by_year.setdefault(year, []).append({
            "name": name,
            "bga_id": bga_id,
            "rank": pos if isinstance(pos, int) else None,
            "note": opm,
        })

    return editions, notes_by_year, be_by_year


def main():
    editions, notes_by_year, be_by_year = load_workbook_data()

    conn = duckdb.connect(str(DB_PATH))

    existing = conn.execute(
        "SELECT COUNT(*) FROM tournaments WHERE type = 'WCC'"
    ).fetchone()[0]
    if existing:
        print(f"{existing} WCC tournaments already present; aborting.")
        conn.close()
        return

    max_tp = conn.execute(
        "SELECT COALESCE(MAX(id), 0) FROM tournament_participants"
    ).fetchone()[0]
    next_tp_id = max_tp + 1

    tournament_id = TOURNAMENT_ID_START
    created_players = []
    matched_players = []

    conn.execute("BEGIN")
    for ed in editions:
        year = ed["year"]
        # Merge top 4 + Belgians, dedup by (name, country_code).
        # Belgian entries override rank/note from top-4.
        participants = {}
        for rank, name, dutch_country in ed["top4"]:
            if not name:
                continue
            ccode = country_code(dutch_country)
            participants[(name, ccode)] = {
                "rank": rank, "name": name, "country": ccode,
                "note": None, "bga_id": None,
            }
        for be in be_by_year.get(year, []):
            key = (be["name"], "BE")
            existing_p = participants.get(key)
            if existing_p:
                if be["rank"]:
                    existing_p["rank"] = be["rank"]
                existing_p["note"] = be["note"]
                existing_p["bga_id"] = be["bga_id"]
            else:
                participants[key] = {
                    "rank": be["rank"], "name": be["name"], "country": "BE",
                    "note": be["note"], "bga_id": be["bga_id"],
                }

        name = f"WK {year}"
        notes = notes_by_year.get(year)
        conn.execute(
            """
            INSERT INTO tournaments
                (id, name, type, year, edition, location,
                 participants_count, notes)
            VALUES (?, ?, 'WCC', ?, ?, ?, ?, ?)
            """,
            [
                tournament_id, name, year, str(ed["edition_no"]),
                ed["location"], ed["participants_count"], notes,
            ],
        )
        print(f"[{tournament_id}] {name}  loc={ed['location']}  "
              f"n={ed['participants_count']}  participants={len(participants)}")

        for p in participants.values():
            pid, created = find_or_create_player(
                conn, p["name"], p["country"], p["bga_id"]
            )
            (created_players if created else matched_players).append(
                (pid, p["name"], p["country"])
            )
            conn.execute(
                """
                INSERT INTO tournament_participants
                    (id, tournament_id, player_id, final_rank, note)
                VALUES (?, ?, ?, ?, ?)
                """,
                [next_tp_id, tournament_id, pid, p["rank"], p["note"]],
            )
            next_tp_id += 1

        tournament_id += 1
    conn.execute("COMMIT")

    print(f"\nMatched {len(matched_players)} existing players:")
    for pid, n, c in matched_players:
        print(f"  [{pid}] {n} ({c})")
    print(f"\nCreated {len(created_players)} new players:")
    for pid, n, c in created_players:
        print(f"  [{pid}] {n} ({c})")

    total = conn.execute(
        "SELECT COUNT(*) FROM tournament_participants "
        "WHERE tournament_id >= ? AND tournament_id < ?",
        [TOURNAMENT_ID_START, tournament_id],
    ).fetchone()[0]
    print(f"\nTotal WCC participant rows inserted: {total}")
    conn.close()


if __name__ == "__main__":
    main()
